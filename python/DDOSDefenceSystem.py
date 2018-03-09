
# coding: utf-8

# #### Live Capture Setup

# In[1]:

# Common Imports

#Pandas for creating dataframes
import pandas as pd

#Pyshark to capture packets
import pyshark

# Ploting
import matplotlib.pyplot as plt


# In[4]:

get_ipython().magic('matplotlib inline')


# In[3]:

# Create capture object for live capture
live_capture = pyshark.LiveCapture(interface='wlan1')


# #### Live Cature (Packet count)

# In[130]:

d = dict()

# Capture 50 live packets
for packet in live_capture.sniff_continuously(packet_count=50):  
        try:
            #print(packet.ip.dst)
            p = packet.ip.dst
            if p in d:
                d[p] += 1
            else:
                d[p] = 1            
        except AttributeError:
            pass
            #print('Ipv4 packet does not exist')
print(d)


# #### Live Cature (Time Frame)

# In[131]:

# Capture packets for 5 second. Time frame is 5 second.
live_capture.sniff(timeout=5)
# print(capture)
# print(len(capture))

# Capture destination of the packet at the router and create dictionary of it.
dst_stream = dict()
for i in range(len(live_capture)):
    try:
        #print(capture[i].ip.dst)
        dst = live_capture[i].ip.dst # This can throw attribute error if dst not exist mostly caused by Ipv6 packets
        if dst in dst_stream:
            dst_stream[dst] += 1
        else:
            dst_stream[dst] = 1         
    except AttributeError:
            pass  #print('Ipv4 packet does not exist')


# In[132]:

dst_stream


# In[122]:

type(live_capture)
#capture[0]


# ### Read from capture file

# In[2]:

# Reading packets from pre-captured file
file_cap = pyshark.FileCapture('captures/botnet-sample.pcap')


# In[3]:

# Convert each request to a data frame

#List of dataframes
dfList = []

#file_cap.set_debug()

for i in range(1000):
#for packet in file_cap:
    layerList = []
    for layer in file_cap[i]:
        layer_dict = {key:value for key, value in layer._all_fields.items() if key != ''}
        layerList.append(pd.DataFrame(layer_dict, index=[0]))
    #print(i,' ',end="")
    dfList.append(pd.concat(layerList, axis=1))
    i += 1
#c = pd.concat(dfList)
# except (AttributeError, AssertionError) as e:
#     print("error")
#     #print('error')
#     continue  #print('Ipv4 packet does not exist')


# In[29]:

# Create big table from all small dataframe which are per packet
first = True
for i in range(len(dfList)):
    if first:
        tdf = dfList[i]
        first = False
        #print(tdf.shape)
    else:
        tdf = tdf.append(dfList[i].loc[:,~dfList[i].columns.duplicated()], ignore_index=True


# In[35]:

# Create big table from all small dataframe which are per packet
first = True
for d in dfList:
    if first:
        tdf = d
        first = False
        #print(tdf.shape)
    else:
        tdf = tdf.append(d.loc[:,~d.columns.duplicated()], ignore_index=True)


# In[37]:

tdf.head()


# In[58]:

t = tdf
i = 'ip.dst'
col = 'ip.dst'#t.columns[4]
print(col)
t = t[t[col].notnull()]
#t[[col,i]][t[col].notnull()]
t[['_ws.expert', 'ip.dst']][t['_ws.expert'].notnull()]


# In[44]:

#Get all non null data
# r = tdf
# for c in list(r.columns):
#     print(c,':',r[c][r[c].notnull()].count())


# In[20]:

get_ipython().magic('time')
dst_cap = dict()
for packet in file_cap:  
        try:
            #print(packet.ip.dst)
            dst = packet.ip.dst
            if dst in dst_cap:
                dst_cap[dst] += 1
            else:
                dst_cap[dst] = 1            
        except AttributeError:
            pass #print('Ipv4 packet does not exist')


# In[57]:

#dst_cap


# In[113]:

dst_cap[]


# #### Create DataFrame

# In[24]:

# Create Dataframe from Dictionary
df = pd.DataFrame.from_dict(dst_cap, orient="index")
df.columns = ['Fequency']


# In[65]:

# Write newly created DataFrame to excel.
from openpyxl import load_workbook
workbook_name = 'data/Attack_Data.xlsx'
book = load_workbook(workbook_name)

# Get last sheet and create new sheet name
last_sheet = book.sheetnames[-1].split("_")
last_sheet_name = last_sheet[0]
last_sheet_no = int(last_sheet[1])
new_sheet = last_sheet_name + '_' + str(last_sheet_no + 1)

# Create a Pandas Excel writer using XlsxWriter as the engine.
#writer = pd.ExcelWriter(workbook_name, engine='xlsxwriter')
writer = pd.ExcelWriter(workbook_name, engine='openpyxl')

#Write DataFrame to excel preserve other sheets
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
df.to_excel(writer, sheet_name=new_sheet, columns=['Fequency'])
writer.save()


# In[1]:

# import subprocess
# subprocess.run(["sudo", "airmon-ng"], stdout=subprocess.PIPE)


# In[76]:

# Load data into dataframe
df = pd.read_excel(workbook_name, sheetname=1)


# In[77]:

df.head()


# In[78]:

df['prob'] = df / df.sum()


# In[81]:

df.sum()


# In[80]:

df.describe()


# In[88]:

df.head()


# In[95]:

df['count'].plot()


# In[104]:

from scipy.stats import entropy
sp.stats.entropy(df['prob'])


# In[178]:

df.sort_values('prob', ascending=False).head(10)


# In[3]:

df1 = pd.read_csv('captures/botnet-capture-20110810-neris.csv')


# In[4]:

df1.head()


# In[162]:

df1.Protocol.unique()


# In[171]:

df2 = df1.groupby(['Destination', 'Protocol'], sort=False).sum()


# In[180]:

df2.sort_values('No.', ascending=False)


# In[ ]:



